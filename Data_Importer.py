from time import strftime
import pandas as pd
import os
import glob
import datetime
import shutil
from openpyxl import load_workbook

path = os.getcwd() + "/Source Files"
csv_files = glob.glob(os.path.join(path, "*.csv"))

destinationName = "Scouting" #Change Source File Name Here
destinationPath = "./" + destinationName + ".xlsx"
sheetName = "Sheet1"    #Change Source Sheet Name Here

shutil.copyfile(destinationPath, "./Backup/" + destinationName + "_" + datetime.datetime.now().strftime("%d-%b-%Y-%H%M%S)") + ".xlsx")

for file in csv_files:

    data = pd.read_csv(file)
    df = pd.DataFrame(data)

    wb = load_workbook(destinationPath)
    wbSheet = wb["Sheet1"]

    with pd.ExcelWriter(destinationPath, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name=sheetName, startrow=wbSheet.max_row, header=False, index=False)

    pd.ExcelWriter.close
    os.remove(file)


    
